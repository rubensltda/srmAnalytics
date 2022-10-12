import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


sheetname='transaction_level'
filename = "qrm_market_value_and_exposure_derivatives_2021-12-31.xlsx"
filepath = "./data/"  + filename

df_transactions = pd.read_excel(filepath, sheet_name=sheetname, skiprows=8)  

##### Data processing #####
# transactions with empty 'Security Identifier', replaced by 'transaction id'
df_transactions['Security Identifier'].fillna(df_transactions["Transaction ID"], inplace=True)
# need to identify products that are in different columns
##### End of Data processing #####


list_allocations = df_transactions['Level 03.1'].unique()
list_indexes = df_transactions['Level 03.3'].unique()
list_products = df_transactions['Level 04'].unique()

print(list_allocations)
print(list_indexes)
print(list_products)


wb = Workbook()

ws = wb.active
ws.title = "transaction_detail"
for r in dataframe_to_rows(df_transactions, index=False, header=True):
    ws.append(r)

ws2 = wb.create_sheet("indices", 0)

ws2.cell(column=1, row=1).value = "Allocations"
row_count = 2
for alloc in list_allocations:
    ws2.cell(column=1, row=row_count).value = alloc
    row_count += 1

ws2.cell(column=3, row=1).value = "Indices"
row_count = 2
for alloc in list_indexes:
    ws2.cell(column=3, row=row_count).value = alloc
    row_count += 1

ws2.cell(column=5, row=1).value = "Products"
row_count = 2
for alloc in list_products:
    ws2.cell(column=5, row=row_count).value = alloc
    row_count += 1


#print(type(list_allocations))

wb.save(filename="./output/validation.xlsx")