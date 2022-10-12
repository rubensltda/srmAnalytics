import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


sheetname='transaction_level'
filename = "qrm_market_value_and_exposure_derivatives_2021-12-31.xlsx"
filepath = "./data/"  + filename

df_transactions = pd.read_excel(filepath, sheet_name=sheetname, skiprows=8)  


##### Data processing #####
# transactions with empty 'Security Identifier', replaced by 'transaction id'
df_transactions['Security Identifier'].fillna(df_transactions["Transaction ID"], inplace=True)

#copying allocation and index columns
df_transactions['df_allocation'] = df_transactions['Level 03.1']
df_transactions['df_index'] = df_transactions['Level 03.3']

# need to identify products that are in different columns
df_transactions['df_product'] = ''
not_aggregate_list = ['ABS MBS','Bond HTM','Bond HTM Accreting','Bond Trading','Bond Trading Accreting','Bond Trading Accreting Callable','Bond Trading Custom','CDs','CPs','Corporates','Deposit','EMP','Governments']

for index, row in df_transactions.iterrows():
    product_found = 'not found'
    
    if str(row['Level 08']) != 'nan' and row['Level 08'] not in not_aggregate_list:
        product_found = row['Level 08']
    elif  str(row['Level 07']) != 'nan' and row['Level 07'] not in not_aggregate_list:
         product_found = row['Level 07']
    elif  str(row['Level 06']) != 'nan' and row['Level 06'] not in not_aggregate_list:
        product_found = row['Level 06']
    elif  str(row['Level 05']) != 'nan' and row['Level 05'] not in not_aggregate_list:
        product_found = row['Level 05']
    else:
        product_found = 'product not found'

    df_transactions.at[index,'df_product'] = product_found
##### End of Data processing #####

list_allocations = df_transactions['df_allocation'].unique()
list_allocations = np.sort(list_allocations)
print(list_allocations)

list_products = df_transactions['df_product'].unique()
list_products = np.sort(list_products)
print(list_products)

list_indexes = df_transactions['df_index'].unique()
list_indexes = np.sort(list_indexes)
print(list_indexes)






###################################################################
# save transaction details
###################################################################
wb = Workbook()
ws = wb.active
#ws.showGridLines = False  ## not working ##

ws.title = "transaction_detail"
for r in dataframe_to_rows(df_transactions, index=False, header=True):
    ws.append(r)
###################################################################


###################################################################
# save unique values for allocation, product, index
###################################################################
ws2 = wb.create_sheet("indices", 0)

ws2.cell(column=1, row=1).value = "Allocations"
row_count = 2
for alloc in list_allocations:
    ws2.cell(column=1, row=row_count).value = alloc
    row_count += 1

ws2.cell(column=3, row=1).value = "Products"
row_count = 2
for product in list_products:
    ws2.cell(column=3, row=row_count).value = product
    row_count += 1

ws2.cell(column=5, row=1).value = "Indices"
row_count = 2
for index in list_indexes:
    ws2.cell(column=5, row=row_count).value = index
    row_count += 1

###################################################################


###################################################################
# save unique values for allocation, product, index
###################################################################
ws3 = wb.create_sheet("exposures", 0)

ws3.cell(column=1, row=1).value = "key_mapping"
ws3.cell(column=2, row=1).value = "Allocation"
ws3.cell(column=3, row=1).value = "Product"
ws3.cell(column=4, row=1).value = "Index"
ws3.cell(column=5, row=1).value = "Face amount"
ws3.cell(column=6, row=1).value = "Book Value"
ws3.cell(column=7, row=1).value = "Market Value"
ws3.cell(column=8, row=1).value = "Economic Value"
ws3.cell(column=9, row=1).value = "BPV"
ws3.cell(column=10, row=1).value = "Convexity"

row_count = 2
for alloc in list_allocations:
    for product in list_products:
        for index in list_indexes:
            ws3.cell(column=1, row=row_count).value = alloc.replace(" ", "_") +  "." + product.replace(" ", "_") + "." + index.replace(" ", "_")
            ws3.cell(column=2, row=row_count).value = alloc
            ws3.cell(column=3, row=row_count).value = product
            ws3.cell(column=4, row=row_count).value = index

            df_filtered = df_transactions[(df_transactions["df_allocation"]==alloc) & (df_transactions["df_index"]==index) & (df_transactions["df_product"]==product)]
            df_face_amount_sum = df_filtered['Face Amount'].sum(axis = 0, skipna = True)
            df_book_value_sum = df_filtered['Book Value'].sum(axis = 0, skipna = True)
            df_market_value_sum = df_filtered['Market Value'].sum(axis = 0, skipna = True)
            df_economic_value_sum = df_filtered['Economic Value'].sum(axis = 0, skipna = True)
            df_bpv_sum = df_filtered['BPV'].sum(axis = 0, skipna = True)
            df_convexity_sum = df_filtered['Convexity'].sum(axis = 0, skipna = True)

            ws3.cell(column=5, row=row_count).value = df_face_amount_sum
            ws3.cell(column=6, row=row_count).value = df_book_value_sum
            ws3.cell(column=7, row=row_count).value = df_market_value_sum
            ws3.cell(column=8, row=row_count).value = df_economic_value_sum
            ws3.cell(column=9, row=row_count).value = df_bpv_sum
            ws3.cell(column=10, row=row_count).value = df_convexity_sum

            row_count += 1
#print(type(list_allocations))
###################################################################

wb.save(filename="./output/risk_mapping.xlsx")

